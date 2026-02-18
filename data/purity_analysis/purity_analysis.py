#!/usr/bin/env python3
"""
Genetic Purity Analysis for Seed Lot Genotyping Data
=====================================================
Analyzes SNP genotyping data to assess seed purity across samples and genotype groups.
Distinguishes hybrid material (expected high heterozygosity) from inbred lines
(expected low heterozygosity).

Input:  260128_ZMS25508_B04_GT_Purity_data_Claude.xlsx
Output: Figures (PNG), summary CSV, and README + Purity Summary tabs added to the xlsx.

Usage:
    python3 purity_analysis.py /path/to/260128_ZMS25508_B04_GT_Purity_data_Claude.xlsx
"""

import sys
import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.cluster.hierarchy import linkage, dendrogram
from scipy.spatial.distance import squareform
from collections import Counter
import warnings
warnings.filterwarnings("ignore")

# ── Configuration ──────────────────────────────────────────────────────────────
IUPAC_HET = {"R": ("A", "G"), "Y": ("C", "T"), "M": ("A", "C"),
              "K": ("G", "T"), "S": ("C", "G"), "W": ("A", "T")}
HET_CODES = set(IUPAC_HET.keys())
HOMO_CODES = {"A", "T", "C", "G"}
VALID_CALLS = HOMO_CODES | HET_CODES

# Hybrid/inbred classification is DATA-DRIVEN: determined from each group's
# median heterozygosity. Groups with median het > 15% are classified as hybrid
# (expected from a cross); groups with median het < 15% are classified as inbred.
HYBRID_HET_THRESHOLD = 0.15  # Median het above this → group is hybrid

# Thresholds for per-sample purity — different for hybrid vs inbred context
# Hybrids: expect ~25-35% het. Low het = selfed/wrong parent.
# Inbreds: expect <5% het. High het = contamination/outcrossing.
HYBRID_HET_MIN = 0.20       # Hybrid with <20% het → warning
HYBRID_HET_WARN = 0.15      # <15% het → fail (not a true hybrid)
INBRED_HET_WARN = 0.05      # Inbred with >5% het → warning
INBRED_HET_FAIL = 0.10      # Inbred with >10% het → fail

FAIL_RATE_WARN = 0.10       # >10% failed calls → poor DNA quality
CONCORDANCE_WARN = 0.90     # <90% concordance → suspect off-type
CONCORDANCE_FAIL = 0.80     # <80% concordance → likely off-type

plt.rcParams.update({
    "font.family": "sans-serif",
    "font.size": 10,
    "axes.titlesize": 13,
    "axes.labelsize": 11,
    "figure.dpi": 150,
})

STATUS_COLORS = {"PASS": "#4CAF50", "WARNING": "#FFC107", "FAIL": "#F44336"}


# ── Data Loading ───────────────────────────────────────────────────────────────
def load_data(xlsx_path):
    raw = pd.read_excel(xlsx_path, sheet_name="Genotypic data", dtype=str)
    source = pd.read_excel(xlsx_path, sheet_name="Genotype and source list")

    sample_cols = raw.columns[1:].tolist()
    gid_row = raw.iloc[0, 1:].values
    geno_row = raw.iloc[1, 1:].values

    markers = raw.iloc[2:, :].copy()
    markers.columns = ["SNP_ID"] + sample_cols
    markers = markers.set_index("SNP_ID")

    meta = pd.DataFrame({
        "sample": sample_cols,
        "GID": pd.to_numeric(gid_row, errors="coerce"),
        "Genotype": geno_row,
    })
    meta["plate_code"] = meta["sample"].str.split(" - ").str[0]
    meta["well"] = meta["sample"].str.split(" - ").str[1]

    meta = meta.merge(
        source[["GID", "PLOT_CODE", "NAME", "Lote de Producción", "PED"]],
        on="GID", how="left"
    )
    # Use Genotype for grouping (finer resolution than NAME).
    # Genotype distinguishes individual inbred lines within the same NAME group.
    # Fall back to NAME if Genotype is missing.
    meta["group"] = meta["Genotype"].where(meta["Genotype"].notna() & (meta["Genotype"] != ""), meta["NAME"])
    # is_hybrid will be set after computing per-sample het rates (data-driven)
    meta["is_hybrid"] = False  # placeholder, updated in main()

    return markers, meta


# ── Per-Sample Metrics ─────────────────────────────────────────────────────────
def compute_sample_metrics(markers, meta):
    records = []
    for _, row in meta.iterrows():
        calls = markers[row["sample"]].values
        total = len(calls)
        n_het = sum(1 for c in calls if c in HET_CODES)
        n_homo = sum(1 for c in calls if c in HOMO_CODES)
        n_fail = sum(1 for c in calls if c == "failed")
        n_valid = n_het + n_homo

        records.append({
            "sample": row["sample"],
            "group": row["group"],
            "Genotype": row["Genotype"],
            "GID": row["GID"],
            "plate_code": row["plate_code"],
            "well": row["well"],
            "Lote": row.get("Lote de Producción", ""),
            "is_hybrid": row["is_hybrid"],
            "total_markers": total,
            "valid_calls": n_valid,
            "het_calls": n_het,
            "homo_calls": n_homo,
            "failed_calls": n_fail,
            "het_rate": n_het / n_valid if n_valid > 0 else np.nan,
            "call_rate": n_valid / total if total > 0 else 0,
            "fail_rate": n_fail / total if total > 0 else 0,
        })
    return pd.DataFrame(records)


# ── Consensus & Concordance ────────────────────────────────────────────────────
def compute_consensus(markers, meta):
    """Most common homozygous call at each locus per group."""
    consensus = {}
    for grp in meta["group"].unique():
        grp_samples = meta.loc[meta["group"] == grp, "sample"].tolist()
        if len(grp_samples) < 2:
            continue
        grp_data = markers[grp_samples]
        cons = []
        for snp in grp_data.index:
            calls = grp_data.loc[snp].values
            homo_calls = [c for c in calls if c in HOMO_CODES]
            if homo_calls:
                cons.append(Counter(homo_calls).most_common(1)[0][0])
            else:
                # For hybrids: take most common het code as consensus
                het_calls = [c for c in calls if c in HET_CODES]
                if het_calls:
                    cons.append(Counter(het_calls).most_common(1)[0][0])
                else:
                    cons.append(np.nan)
        consensus[grp] = pd.Series(cons, index=grp_data.index)
    return consensus


def compute_concordance(markers, meta, consensus):
    concordances = []
    for _, row in meta.iterrows():
        grp = row["group"]
        if grp not in consensus:
            concordances.append(np.nan)
            continue

        calls = markers[row["sample"]]
        cons = consensus[grp]
        valid_mask = cons.notna() & calls.isin(VALID_CALLS)
        n_valid = valid_mask.sum()
        if n_valid == 0:
            concordances.append(np.nan)
            continue

        n_concordant = 0
        for snp in cons.index[valid_mask]:
            call = calls[snp]
            expected = cons[snp]
            if call == expected:
                n_concordant += 1
            elif call in HET_CODES and expected in HOMO_CODES:
                # Het call containing expected allele = partial concordance
                if expected in IUPAC_HET[call]:
                    n_concordant += 1
            elif call in HOMO_CODES and expected in HET_CODES:
                # Homo call that is one of the expected het alleles
                if call in IUPAC_HET[expected]:
                    n_concordant += 1
        concordances.append(n_concordant / n_valid)
    return concordances


# ── Purity Classification ─────────────────────────────────────────────────────
def classify_purity(metrics):
    statuses = []
    reasons = []
    for _, row in metrics.iterrows():
        status = "PASS"
        parts = []

        # DNA quality check
        if row["fail_rate"] > FAIL_RATE_WARN:
            status = "WARNING"
            parts.append(f"high fail rate ({row['fail_rate']:.1%})")

        # Heterozygosity check — context-dependent
        if row["is_hybrid"]:
            # Hybrids SHOULD be heterozygous. Low het = problem.
            if row["het_rate"] < HYBRID_HET_WARN:
                status = "FAIL"
                parts.append(f"low heterozygosity for hybrid ({row['het_rate']:.1%}; expected >20%)")
            elif row["het_rate"] < HYBRID_HET_MIN:
                if status != "FAIL":
                    status = "WARNING"
                parts.append(f"reduced heterozygosity for hybrid ({row['het_rate']:.1%})")
        else:
            # Inbred lines should have LOW het.
            if row["het_rate"] > INBRED_HET_FAIL:
                status = "FAIL"
                parts.append(f"high heterozygosity for inbred ({row['het_rate']:.1%})")
            elif row["het_rate"] > INBRED_HET_WARN:
                if status != "FAIL":
                    status = "WARNING"
                parts.append(f"elevated heterozygosity for inbred ({row['het_rate']:.1%})")

        # Concordance check
        if pd.notna(row.get("concordance")):
            if row["concordance"] < CONCORDANCE_FAIL:
                status = "FAIL"
                parts.append(f"low concordance ({row['concordance']:.1%})")
            elif row["concordance"] < CONCORDANCE_WARN:
                if status != "FAIL":
                    status = "WARNING"
                parts.append(f"reduced concordance ({row['concordance']:.1%})")

        statuses.append(status)
        reasons.append("; ".join(parts) if parts else "OK")

    metrics["purity_status"] = statuses
    metrics["purity_reason"] = reasons
    return metrics


# ── IBS Distance ───────────────────────────────────────────────────────────────
def compute_ibs_distance(markers, sample_list):
    allele_map = {"A": 0, "C": 1, "G": 2, "T": 3}
    n_markers = len(markers.index)
    n_samples = len(sample_list)

    mat = np.full((n_markers, n_samples), np.nan)
    for j, s in enumerate(sample_list):
        col = markers[s].values
        for i, c in enumerate(col):
            if c in allele_map:
                mat[i, j] = allele_map[c]
            elif c in IUPAC_HET:
                a1, a2 = IUPAC_HET[c]
                mat[i, j] = (allele_map[a1] + allele_map[a2]) / 2.0

    dist_mat = np.zeros((n_samples, n_samples))
    for i in range(n_samples):
        for j in range(i + 1, n_samples):
            valid = ~np.isnan(mat[:, i]) & ~np.isnan(mat[:, j])
            if valid.sum() > 0:
                diff = np.sum(mat[valid, i] != mat[valid, j]) / valid.sum()
            else:
                diff = 1.0
            dist_mat[i, j] = diff
            dist_mat[j, i] = diff
    return dist_mat


# ── Figures ────────────────────────────────────────────────────────────────────
def fig1_heterozygosity_by_group(metrics, outdir):
    """Boxplot of heterozygosity with hybrid/inbred context."""
    fig, axes = plt.subplots(1, 2, figsize=(16, 6), gridspec_kw={"width_ratios": [3, 2]})

    # Left: Hybrids
    ax = axes[0]
    hybrid_m = metrics[metrics["is_hybrid"]]
    if len(hybrid_m) > 0:
        groups = hybrid_m.groupby("group")["het_rate"].median().sort_values(ascending=False).index
        data = [hybrid_m.loc[hybrid_m["group"] == g, "het_rate"].values for g in groups]
        bp = ax.boxplot(data, labels=groups, patch_artist=True, showfliers=True,
                        flierprops=dict(marker="o", markersize=3, alpha=0.5))
        colors = plt.cm.Blues(np.linspace(0.3, 0.8, len(groups)))
        for patch, color in zip(bp["boxes"], colors):
            patch.set_facecolor(color)
            patch.set_alpha(0.8)
        ax.axhline(HYBRID_HET_MIN, color="orange", ls="--", alpha=0.7,
                    label=f"Warning (<{HYBRID_HET_MIN:.0%})")
        ax.axhline(HYBRID_HET_WARN, color="red", ls="--", alpha=0.7,
                    label=f"Fail (<{HYBRID_HET_WARN:.0%})")
    ax.set_title("Hybrid Material\n(high heterozygosity expected)")
    ax.set_ylabel("Heterozygosity Rate")
    ax.legend(fontsize=8)
    ax.tick_params(axis="x", rotation=45)

    # Right: Inbred
    ax = axes[1]
    inbred_m = metrics[~metrics["is_hybrid"]]
    if len(inbred_m) > 0:
        groups = inbred_m.groupby("group")["het_rate"].median().sort_values(ascending=False).index
        data = [inbred_m.loc[inbred_m["group"] == g, "het_rate"].values for g in groups]
        bp = ax.boxplot(data, labels=groups, patch_artist=True, showfliers=True,
                        flierprops=dict(marker="o", markersize=3, alpha=0.5))
        colors = plt.cm.Oranges(np.linspace(0.3, 0.8, len(groups)))
        for patch, color in zip(bp["boxes"], colors):
            patch.set_facecolor(color)
            patch.set_alpha(0.8)
        ax.axhline(INBRED_HET_WARN, color="orange", ls="--", alpha=0.7,
                    label=f"Warning (>{INBRED_HET_WARN:.0%})")
        ax.axhline(INBRED_HET_FAIL, color="red", ls="--", alpha=0.7,
                    label=f"Fail (>{INBRED_HET_FAIL:.0%})")
    ax.set_title("Inbred Lines\n(low heterozygosity expected)")
    ax.legend(fontsize=8)
    ax.tick_params(axis="x", rotation=45)

    fig.suptitle("Heterozygosity Rate by Genotype Group", fontsize=14, y=1.02)
    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig1_heterozygosity_by_group.png"), bbox_inches="tight")
    plt.close()


def fig2_concordance_by_group(metrics, outdir):
    fig, ax = plt.subplots(figsize=(14, 6))
    valid = metrics.dropna(subset=["concordance"])
    groups_ordered = valid.groupby("group")["concordance"].median().sort_values().index
    data = [valid.loc[valid["group"] == g, "concordance"].values for g in groups_ordered]

    bp = ax.boxplot(data, labels=groups_ordered, patch_artist=True, showfliers=True,
                    flierprops=dict(marker="o", markersize=4, alpha=0.6))
    # Color by hybrid/inbred
    hyb_set = set(metrics.loc[metrics["is_hybrid"], "group"].unique())
    for i, grp in enumerate(groups_ordered):
        color = "#5C9BD1" if grp in hyb_set else "#E8915A"
        bp["boxes"][i].set_facecolor(color)
        bp["boxes"][i].set_alpha(0.7)

    ax.axhline(CONCORDANCE_WARN, color="orange", ls="--", alpha=0.7, label=f"Warning ({CONCORDANCE_WARN:.0%})")
    ax.axhline(CONCORDANCE_FAIL, color="red", ls="--", alpha=0.7, label=f"Fail ({CONCORDANCE_FAIL:.0%})")
    ax.set_ylabel("Concordance with Group Consensus")
    ax.set_title("Sample-to-Consensus Concordance by Group\n(Blue = Hybrid, Orange = Inbred)")
    ax.legend(loc="lower left")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig2_concordance_by_group.png"), bbox_inches="tight")
    plt.close()


def fig3_purity_summary(metrics, outdir):
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))

    # Left: stacked bar per group
    ct = pd.crosstab(metrics["group"], metrics["purity_status"])
    for col in ["PASS", "WARNING", "FAIL"]:
        if col not in ct.columns:
            ct[col] = 0
    ct = ct[["PASS", "WARNING", "FAIL"]]
    ct = ct.loc[ct.sum(axis=1).sort_values(ascending=False).index]
    ct.plot(kind="bar", stacked=True, ax=ax1, color=[STATUS_COLORS[s] for s in ["PASS", "WARNING", "FAIL"]])
    ax1.set_ylabel("Number of Samples")
    ax1.set_title("Purity Classification by Group")
    ax1.legend(title="Status")
    ax1.tick_params(axis="x", rotation=45)

    # Right: overall pie
    counts = metrics["purity_status"].value_counts()
    ax2.pie([counts.get(s, 0) for s in ["PASS", "WARNING", "FAIL"]],
            labels=["PASS", "WARNING", "FAIL"],
            colors=[STATUS_COLORS[s] for s in ["PASS", "WARNING", "FAIL"]],
            autopct="%1.0f%%", startangle=90, textprops={"fontsize": 12})
    ax2.set_title(f"Overall Purity (n={len(metrics)})")

    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig3_purity_summary.png"), bbox_inches="tight")
    plt.close()


def fig4_scatter_het_concordance(metrics, outdir):
    fig, ax = plt.subplots(figsize=(12, 8))
    valid = metrics.dropna(subset=["concordance"])

    for status in ["PASS", "WARNING", "FAIL"]:
        sub = valid[valid["purity_status"] == status]
        marker = "o" if not sub.empty and sub["is_hybrid"].iloc[0] else "s"
        ax.scatter(sub["het_rate"], sub["concordance"], c=STATUS_COLORS[status],
                   label=f"{status} (n={len(sub)})", alpha=0.5, s=25, edgecolors="none")

    # Add group labels for outliers
    outliers = valid[(valid["concordance"] < CONCORDANCE_WARN) | (valid["het_rate"] > 0.4)]
    for _, row in outliers.iterrows():
        ax.annotate(row["group"], (row["het_rate"], row["concordance"]),
                    fontsize=6, alpha=0.7)

    ax.axhline(CONCORDANCE_WARN, color="orange", ls="--", alpha=0.4, linewidth=0.8)
    ax.axhline(CONCORDANCE_FAIL, color="red", ls="--", alpha=0.4, linewidth=0.8)
    ax.axvline(HYBRID_HET_MIN, color="blue", ls=":", alpha=0.3, linewidth=0.8, label="Hybrid het minimum")
    ax.axvline(INBRED_HET_FAIL, color="red", ls=":", alpha=0.3, linewidth=0.8, label="Inbred het fail")

    ax.set_xlabel("Heterozygosity Rate")
    ax.set_ylabel("Concordance with Group Consensus")
    ax.set_title("Heterozygosity vs. Concordance (colored by purity status)")
    ax.legend(fontsize=8, loc="lower left")
    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig4_het_vs_concordance.png"), bbox_inches="tight")
    plt.close()


def fig5_dendrogram(markers, meta, outdir, top_n=6, samples_per_group=15):
    """Hierarchical clustering of samples from the largest groups."""
    top_groups = meta["group"].value_counts().head(top_n).index.tolist()
    selected = []
    labels = []
    for grp in top_groups:
        grp_samples = meta.loc[meta["group"] == grp, "sample"].tolist()
        subset = grp_samples[:samples_per_group]
        selected.extend(subset)
        labels.extend([grp] * len(subset))

    if len(selected) < 4:
        return

    print(f"  Computing IBS distances for {len(selected)} samples...")
    dist_mat = compute_ibs_distance(markers, selected)
    condensed = squareform(dist_mat)
    Z = linkage(condensed, method="ward")

    fig, ax = plt.subplots(figsize=(18, 7))
    unique_groups = list(dict.fromkeys(labels))
    group_cmap = plt.cm.tab10(np.linspace(0, 1, len(unique_groups)))
    color_map = {g: matplotlib.colors.rgb2hex(group_cmap[i]) for i, g in enumerate(unique_groups)}

    dendrogram(Z, labels=labels, ax=ax, leaf_rotation=90,
               leaf_font_size=6, link_color_func=lambda k: "#aaaaaa")

    ax.set_title(f"Hierarchical Clustering — Top {top_n} Groups ({samples_per_group} samples each)")
    ax.set_ylabel("Distance")

    # Add color legend
    from matplotlib.patches import Patch
    legend_elements = [Patch(facecolor=color_map[g], label=g) for g in unique_groups]
    ax.legend(handles=legend_elements, loc="upper right", fontsize=8)

    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig5_dendrogram.png"), bbox_inches="tight")
    plt.close()


def fig6_fail_rate(metrics, outdir):
    fig, ax = plt.subplots(figsize=(14, 5))
    group_fail = metrics.groupby("group")["fail_rate"].agg(["mean", "std"])
    group_fail = group_fail.sort_values("mean", ascending=False)

    hyb_set = set(metrics.loc[metrics["is_hybrid"], "group"].unique())
    colors = ["#5C9BD1" if g in hyb_set else "#E8915A" for g in group_fail.index]
    ax.bar(range(len(group_fail)), group_fail["mean"], yerr=group_fail["std"],
           capsize=2, color=colors, alpha=0.8)
    ax.set_xticks(range(len(group_fail)))
    ax.set_xticklabels(group_fail.index, rotation=45, ha="right")
    ax.axhline(FAIL_RATE_WARN, color="red", ls="--", alpha=0.7, label=f"Warning ({FAIL_RATE_WARN:.0%})")
    ax.set_ylabel("Mean Fail Rate")
    ax.set_title("Genotyping Failure Rate by Group (mean ± std)\n(Blue = Hybrid, Orange = Inbred)")
    ax.legend()
    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig6_fail_rate.png"), bbox_inches="tight")
    plt.close()


def fig7_group_detail_grid(metrics, outdir):
    """Per-group summary dashboard: het rate distribution + concordance + status counts."""
    groups = metrics["group"].value_counts().index.tolist()
    n_groups = len(groups)
    fig, axes = plt.subplots(n_groups, 3, figsize=(16, 3 * n_groups))
    if n_groups == 1:
        axes = axes.reshape(1, -1)

    hyb_set = set(metrics.loc[metrics["is_hybrid"], "group"].unique())
    for i, grp in enumerate(groups):
        sub = metrics[metrics["group"] == grp]
        is_hyb = grp in hyb_set
        label = f"{grp} ({'hybrid' if is_hyb else 'inbred'}, n={len(sub)})"

        # Col 1: Het rate histogram
        ax = axes[i, 0]
        ax.hist(sub["het_rate"], bins=30, color="#5C9BD1" if is_hyb else "#E8915A", alpha=0.8, edgecolor="white")
        if is_hyb:
            ax.axvline(HYBRID_HET_MIN, color="orange", ls="--", alpha=0.7)
        else:
            ax.axvline(INBRED_HET_WARN, color="orange", ls="--", alpha=0.7)
            ax.axvline(INBRED_HET_FAIL, color="red", ls="--", alpha=0.7)
        ax.set_ylabel(label, fontsize=9, fontweight="bold")
        if i == 0:
            ax.set_title("Heterozygosity Rate")

        # Col 2: Concordance histogram
        ax = axes[i, 1]
        conc = sub["concordance"].dropna()
        if len(conc) > 0:
            ax.hist(conc, bins=30, color="#7E8CE0", alpha=0.8, edgecolor="white")
            ax.axvline(CONCORDANCE_WARN, color="orange", ls="--", alpha=0.7)
            ax.axvline(CONCORDANCE_FAIL, color="red", ls="--", alpha=0.7)
        if i == 0:
            ax.set_title("Concordance")

        # Col 3: Status pie
        ax = axes[i, 2]
        counts = sub["purity_status"].value_counts()
        vals = [counts.get(s, 0) for s in ["PASS", "WARNING", "FAIL"]]
        if sum(vals) > 0:
            ax.pie(vals, labels=["PASS", "WARNING", "FAIL"],
                   colors=[STATUS_COLORS[s] for s in ["PASS", "WARNING", "FAIL"]],
                   autopct=lambda p: f"{p:.0f}%" if p > 0 else "", startangle=90, textprops={"fontsize": 8})
        if i == 0:
            ax.set_title("Purity Status")

    fig.suptitle("Per-Group Purity Detail", fontsize=14, y=1.01)
    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig7_group_detail.png"), bbox_inches="tight")
    plt.close()


# ── README Tab ─────────────────────────────────────────────────────────────────
def build_readme_text(metrics):
    n_samples = len(metrics)
    n_pass = (metrics["purity_status"] == "PASS").sum()
    n_warn = (metrics["purity_status"] == "WARNING").sum()
    n_fail = (metrics["purity_status"] == "FAIL").sum()
    n_groups = metrics["group"].nunique()
    n_markers = metrics["total_markers"].iloc[0]
    n_hybrid = metrics[metrics["is_hybrid"]].shape[0]
    n_inbred = n_samples - n_hybrid

    # Group-level summary
    hyb_set = set(metrics.loc[metrics["is_hybrid"], "group"].unique())
    group_lines = []
    for grp in metrics["group"].value_counts().index:
        sub = metrics[metrics["group"] == grp]
        is_hyb = grp in hyb_set
        n_p = (sub["purity_status"] == "PASS").sum()
        n_w = (sub["purity_status"] == "WARNING").sum()
        n_f = (sub["purity_status"] == "FAIL").sum()
        med_het = sub["het_rate"].median()
        med_conc = sub["concordance"].median()
        group_lines.append(
            f"  {grp:25s}  {'hybrid' if is_hyb else 'inbred':7s}  "
            f"n={len(sub):3d}  PASS={n_p:3d}  WARN={n_w:3d}  FAIL={n_f:3d}  "
            f"med_het={med_het:.2%}  med_conc={med_conc:.2%}"
        )

    readme = f"""README — Genetic Purity Analysis (ZMS25508 B04)
=================================================

DATA OVERVIEW
This workbook contains SNP genotyping data for a seed purity assessment. The "Genotypic data" sheet has {n_markers:,} SNP markers (rows, Affymetrix probe IDs like AX-XXXXXXXXX) genotyped across {n_samples} individual seed samples (columns). Each column header encodes a plate code and well position (e.g., "F6601 - P1A1"). The first two data rows contain metadata: the GID (germplasm identifier) and Genotype name for each sample.

Genotype calls use IUPAC codes: A, T, C, G for homozygous calls; R (A/G), Y (C/T), M (A/C), K (G/T), S (C/G), W (A/T) for heterozygous calls. Failed assays are marked "failed".

The "Genotype and source list" sheet maps each GID to its genotype name, pedigree, seed lot ("Lote de Produccion"), and accession info. Samples belong to {n_groups} genotype groups: {n_hybrid} samples are hybrid material (expected heterozygous) and {n_inbred} are inbred lines (expected homozygous).

HYBRID vs. INBRED CONTEXT
This is critical: hybrid seed (F5, F5-DOBLE, Hembra F5, etc.) is EXPECTED to have ~30-35% heterozygosity because it results from crossing two different inbred parents. Low heterozygosity in a hybrid indicates selfing or wrong parentage. Conversely, inbred lines (CLWQHZN86, etc.) should be nearly all homozygous; high heterozygosity indicates contamination or outcrossing.

ANALYSIS METHODOLOGY
Three metrics assess purity:
1. HETEROZYGOSITY RATE: Proportion of het calls among valid calls. Evaluated against group-appropriate thresholds:
   - Hybrids: <20% = warning, <15% = fail (too low)
   - Inbreds: >5% = warning, >10% = fail (too high)
2. GENOTYPING FAIL RATE: >10% failed markers = warning (poor DNA quality)
3. CONCORDANCE WITH GROUP CONSENSUS: Each sample compared to the most common genotype at each locus within its group. <90% = warning, <80% = fail (potential off-type or mislabeling)

RESULTS SUMMARY
  PASS:    {n_pass:4d} samples ({n_pass/n_samples:.1%})
  WARNING: {n_warn:4d} samples ({n_warn/n_samples:.1%})
  FAIL:    {n_fail:4d} samples ({n_fail/n_samples:.1%})

PER-GROUP BREAKDOWN:
{chr(10).join(group_lines)}

FIGURES (saved alongside this workbook):
  fig1: Heterozygosity boxplots split by hybrid/inbred context
  fig2: Concordance boxplots by group
  fig3: Purity classification stacked bar + overall pie
  fig4: Heterozygosity vs. concordance scatter
  fig5: Hierarchical clustering dendrogram (top groups)
  fig6: Genotyping failure rate by group
  fig7: Per-group detail dashboard (het, concordance, status)

HOW TO REPRODUCE:
  python3 purity_analysis.py <path_to_this_xlsx>
  Requires: pandas, numpy, matplotlib, scipy, openpyxl
"""
    return readme


def add_readme_tab(xlsx_path, readme_text, metrics):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font

    wb = load_workbook(xlsx_path)
    for name in ["README", "Purity Summary"]:
        if name in wb.sheetnames:
            del wb[name]

    # README tab
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 130
    ws["A1"] = readme_text
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A1"].font = Font(name="Consolas", size=10)
    ws.row_dimensions[1].height = 900

    # Purity Summary tab
    ws2 = wb.create_sheet("Purity Summary", 1)
    cols = ["sample", "group", "Genotype", "is_hybrid", "Lote", "het_rate", "call_rate",
            "fail_rate", "concordance", "purity_status", "purity_reason"]
    available = [c for c in cols if c in metrics.columns]
    summary = metrics[available].sort_values(["purity_status", "group", "het_rate"],
                                              ascending=[True, True, False])

    for j, col in enumerate(available, 1):
        cell = ws2.cell(row=1, column=j, value=col)
        cell.font = Font(bold=True)

    for i, (_, row) in enumerate(summary.iterrows(), 2):
        for j, col in enumerate(available, 1):
            val = row[col]
            if isinstance(val, float) and col in ("het_rate", "call_rate", "fail_rate", "concordance"):
                ws2.cell(row=i, column=j, value=round(val, 4))
            elif isinstance(val, (bool, np.bool_)):
                ws2.cell(row=i, column=j, value=str(val))
            else:
                ws2.cell(row=i, column=j, value=val)

    for j, col in enumerate(available, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=j).column_letter].width = max(len(col) + 4, 15)

    wb.save(xlsx_path)
    print(f"  Added README and Purity Summary tabs to workbook")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        # Try to find the file automatically
        candidates = [
            os.path.expanduser("~/personal_website/Genetic purity analysis/260128_ZMS25508_B04_GT_Purity_data_Claude.xlsx"),
        ]
        xlsx_path = None
        for c in candidates:
            if os.path.exists(c):
                xlsx_path = c
                break
        if xlsx_path is None:
            print("Usage: python3 purity_analysis.py <path_to_xlsx>")
            sys.exit(1)
    else:
        xlsx_path = sys.argv[1]

    if not os.path.exists(xlsx_path):
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    outdir = os.path.dirname(os.path.abspath(__file__))
    print(f"Input:  {xlsx_path}")
    print(f"Output: {outdir}")
    print()

    print("Loading data...")
    markers, meta = load_data(xlsx_path)
    print(f"  {len(markers)} markers x {len(meta)} samples, {meta['group'].nunique()} groups")

    print("Computing per-sample metrics...")
    metrics = compute_sample_metrics(markers, meta)

    # Data-driven hybrid classification: groups with median het > threshold are hybrid
    print("Classifying hybrid vs inbred groups (data-driven)...")
    group_med_het = metrics.groupby("group")["het_rate"].median()
    hybrid_groups = set(group_med_het[group_med_het > HYBRID_HET_THRESHOLD].index)
    metrics["is_hybrid"] = metrics["group"].isin(hybrid_groups)
    meta["is_hybrid"] = meta["group"].isin(hybrid_groups)
    for grp in sorted(hybrid_groups):
        print(f"  HYBRID: {grp:25s} (median het = {group_med_het[grp]:.2%})")
    inbred_groups = set(group_med_het.index) - hybrid_groups
    for grp in sorted(inbred_groups):
        print(f"  INBRED: {grp:25s} (median het = {group_med_het[grp]:.2%})")
    n_hyb = metrics["is_hybrid"].sum()
    n_inb = (~metrics["is_hybrid"]).sum()
    print(f"  Total: {n_hyb} hybrid samples, {n_inb} inbred samples")

    print("Computing group consensus genotypes...")
    consensus = compute_consensus(markers, meta)

    print("Computing concordance...")
    metrics["concordance"] = compute_concordance(markers, meta, consensus)

    print("Classifying purity...")
    metrics = classify_purity(metrics)

    # Summary
    print()
    print("=" * 65)
    print("PURITY SUMMARY")
    print("=" * 65)
    for status in ["PASS", "WARNING", "FAIL"]:
        n = (metrics["purity_status"] == status).sum()
        print(f"  {status:8s}: {n:4d} samples ({n/len(metrics):.1%})")
    print()

    # Per-group summary
    print("PER-GROUP BREAKDOWN:")
    for grp in metrics["group"].value_counts().index:
        sub = metrics[metrics["group"] == grp]
        is_hyb = sub["is_hybrid"].iloc[0] if len(sub) > 0 else False
        n_p = (sub["purity_status"] == "PASS").sum()
        n_w = (sub["purity_status"] == "WARNING").sum()
        n_f = (sub["purity_status"] == "FAIL").sum()
        print(f"  {grp:25s} ({'hybrid' if is_hyb else 'inbred':7s})  "
              f"n={len(sub):3d}  PASS={n_p:3d}  WARN={n_w:3d}  FAIL={n_f:3d}  "
              f"med_het={sub['het_rate'].median():.2%}  med_conc={sub['concordance'].median():.2%}")
    print()

    # Flag problems
    fails = metrics[metrics["purity_status"] == "FAIL"]
    if len(fails) > 0:
        print(f"FAILED SAMPLES ({len(fails)}):")
        for _, row in fails.head(30).iterrows():
            conc_str = f"{row['concordance']:.2%}" if pd.notna(row['concordance']) else "N/A"
            print(f"  {row['sample']:25s}  {row['group']:20s}  het={row['het_rate']:.2%}  "
                  f"conc={conc_str}  {row['purity_reason']}")
        if len(fails) > 30:
            print(f"  ... and {len(fails) - 30} more (see purity_summary.csv)")
    print()

    # Figures
    print("Generating figures...")
    fig1_heterozygosity_by_group(metrics, outdir)
    print("  fig1_heterozygosity_by_group.png")
    fig2_concordance_by_group(metrics, outdir)
    print("  fig2_concordance_by_group.png")
    fig3_purity_summary(metrics, outdir)
    print("  fig3_purity_summary.png")
    fig4_scatter_het_concordance(metrics, outdir)
    print("  fig4_het_vs_concordance.png")
    fig5_dendrogram(markers, meta, outdir)
    print("  fig5_dendrogram.png")
    fig6_fail_rate(metrics, outdir)
    print("  fig6_fail_rate.png")
    fig7_group_detail_grid(metrics, outdir)
    print("  fig7_group_detail.png")

    # CSV
    csv_path = os.path.join(outdir, "purity_summary.csv")
    metrics.to_csv(csv_path, index=False)
    print(f"\nSaved: {csv_path}")

    # Add tabs to xlsx
    print("\nUpdating workbook...")
    readme_text = build_readme_text(metrics)
    add_readme_tab(xlsx_path, readme_text, metrics)

    print("\nDone!")


if __name__ == "__main__":
    main()
