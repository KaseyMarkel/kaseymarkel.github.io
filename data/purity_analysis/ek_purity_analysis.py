#!/usr/bin/env python3
"""
EK Purity Analysis — Companion analysis of the breeding director's manual
classification of F5/F5-DOBLE hybrid samples from ZMS25508 B04.

This spreadsheet covers the 273 hybrid samples (a subset of the full 538)
and adds distance-based classification: Self parent, Out-type/outcross,
and distance metrics to F5 and Hembra consensus genotypes.

Input:  260128_ZMS25508_B04_GT_EK_F5 purity_12_02_2026.xlsx
Output: Figures (PNG), summary CSV, README tab added to xlsx.

Usage:
    python3 ek_purity_analysis.py /path/to/EK_xlsx [/path/to/original_purity_summary.csv]
"""

import sys
import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import warnings
warnings.filterwarnings("ignore")

plt.rcParams.update({
    "font.family": "sans-serif",
    "font.size": 10,
    "axes.titlesize": 13,
    "axes.labelsize": 11,
    "figure.dpi": 150,
})

STATUS_COLORS = {
    "Pure hybrid": "#4CAF50",
    "Self parent": "#F44336",
    "Outcross": "#FF9800",
}


# ── Data Loading ───────────────────────────────────────────────────────────────
def load_ek(xlsx_path):
    """Load the EK purity spreadsheet."""
    df = pd.read_excel(xlsx_path, sheet_name="Pureza de F5")
    summary = pd.read_excel(xlsx_path, sheet_name="Quick summary")

    # Clean up
    df["het_rate"] = pd.to_numeric(df["% HETERO"], errors="coerce") / 100.0
    df["dist_F5"] = pd.to_numeric(df["dist_prom_F5"], errors="coerce")
    df["dist_hembra"] = pd.to_numeric(df["dist_prom_hembra"], errors="coerce")
    df["closer_to_hembra"] = df["mas_cerca_de_hembra"].astype(bool) if df["mas_cerca_de_hembra"].notna().any() else False

    # Classification
    df["ek_class"] = "Pure hybrid"
    df.loc[df["Self parent"] == "Yes", "ek_class"] = "Self parent"
    df.loc[df["Out-type/outcross"].notna(), "ek_class"] = "Outcross"

    return df, summary


def load_our_analysis(csv_path):
    """Load our automated analysis results for comparison."""
    if csv_path and os.path.exists(csv_path):
        return pd.read_csv(csv_path)
    return None


# ── Figures ────────────────────────────────────────────────────────────────────
def fig_ek1_distance_scatter(df, outdir):
    """Distance to F5 vs distance to Hembra, colored by EK classification."""
    fig, ax = plt.subplots(figsize=(10, 8))

    for cls in ["Pure hybrid", "Self parent", "Outcross"]:
        sub = df[df["ek_class"] == cls]
        ax.scatter(sub["dist_F5"], sub["dist_hembra"], c=STATUS_COLORS[cls],
                   label=f"{cls} (n={len(sub)})", alpha=0.6, s=30, edgecolors="none")

    ax.set_xlabel("Genetic Distance to F5 Hybrid Consensus")
    ax.set_ylabel("Genetic Distance to Hembra (Female Parent) Consensus")
    ax.set_title("Sample Classification by Genetic Distance\n(EK Manual Analysis)")

    # Add diagonal reference
    lims = [0, max(ax.get_xlim()[1], ax.get_ylim()[1])]
    ax.plot(lims, lims, "k--", alpha=0.2, linewidth=0.8, label="Equal distance")

    ax.legend(fontsize=9)
    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig_ek1_distance_scatter.png"), bbox_inches="tight")
    plt.close()


def fig_ek2_het_by_classification(df, outdir):
    """Heterozygosity distribution by EK classification."""
    fig, axes = plt.subplots(1, 3, figsize=(16, 5), sharey=True)

    for i, cls in enumerate(["Pure hybrid", "Self parent", "Outcross"]):
        ax = axes[i]
        sub = df[df["ek_class"] == cls]
        ax.hist(sub["het_rate"] * 100, bins=25, color=STATUS_COLORS[cls],
                alpha=0.8, edgecolor="white")
        ax.set_xlabel("Heterozygosity (%)")
        ax.set_title(f"{cls}\n(n={len(sub)}, mean={sub['het_rate'].mean()*100:.1f}%)")
        ax.axvline(20, color="red", ls="--", alpha=0.5, linewidth=0.8)
        if i == 0:
            ax.set_ylabel("Number of Samples")

    fig.suptitle("Heterozygosity Distribution by Classification", fontsize=14, y=1.02)
    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig_ek2_het_by_classification.png"), bbox_inches="tight")
    plt.close()


def fig_ek3_lot_purity(df, outdir):
    """Per-lot stacked bar of classification + impurity rates."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))

    # Left: stacked bar per lot
    lots = df["Lote de Producción"].value_counts().index.tolist()
    ct = pd.crosstab(df["Lote de Producción"], df["ek_class"])
    for cls in ["Pure hybrid", "Self parent", "Outcross"]:
        if cls not in ct.columns:
            ct[cls] = 0
    ct = ct[["Pure hybrid", "Self parent", "Outcross"]]
    ct = ct.loc[lots]

    ct.plot(kind="bar", stacked=True, ax=ax1,
            color=[STATUS_COLORS[c] for c in ["Pure hybrid", "Self parent", "Outcross"]])
    ax1.set_ylabel("Number of Samples")
    ax1.set_title("Sample Classification by Seed Lot")
    ax1.legend(title="Classification")
    ax1.tick_params(axis="x", rotation=45)

    # Right: impurity rate per lot
    lot_stats = []
    for lot in lots:
        sub = df[df["Lote de Producción"] == lot]
        n_total = len(sub)
        n_self = (sub["ek_class"] == "Self parent").sum()
        n_out = (sub["ek_class"] == "Outcross").sum()
        lot_stats.append({
            "lot": lot,
            "n": n_total,
            "self_rate": n_self / n_total,
            "outcross_rate": n_out / n_total,
            "total_impurity": (n_self + n_out) / n_total,
        })
    lot_df = pd.DataFrame(lot_stats)

    x = range(len(lot_df))
    w = 0.35
    ax2.bar([i - w/2 for i in x], lot_df["self_rate"] * 100, w,
            color=STATUS_COLORS["Self parent"], alpha=0.8, label="Self parent %")
    ax2.bar([i + w/2 for i in x], lot_df["outcross_rate"] * 100, w,
            color=STATUS_COLORS["Outcross"], alpha=0.8, label="Outcross %")

    # Add total impurity line
    ax2.plot(x, lot_df["total_impurity"] * 100, "ko-", markersize=6,
             label="Total impurity %", linewidth=1.5)

    ax2.set_xticks(x)
    ax2.set_xticklabels(lot_df["lot"], rotation=45, ha="right")
    ax2.set_ylabel("Impurity Rate (%)")
    ax2.set_title("Impurity Rate by Seed Lot")
    ax2.legend(fontsize=9)

    # Add sample count annotations
    for i, row in lot_df.iterrows():
        ax2.annotate(f"n={row['n']:.0f}", (i, row["total_impurity"] * 100 + 1),
                     ha="center", fontsize=8, alpha=0.7)

    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig_ek3_lot_purity.png"), bbox_inches="tight")
    plt.close()

    return lot_df


def fig_ek4_distance_vs_het(df, outdir):
    """Het rate vs distance-to-F5, showing how the metrics correlate."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

    for cls in ["Pure hybrid", "Self parent", "Outcross"]:
        sub = df[df["ek_class"] == cls]
        ax1.scatter(sub["het_rate"] * 100, sub["dist_F5"], c=STATUS_COLORS[cls],
                    label=f"{cls} (n={len(sub)})", alpha=0.6, s=25, edgecolors="none")
        ax2.scatter(sub["het_rate"] * 100, sub["dist_hembra"], c=STATUS_COLORS[cls],
                    label=f"{cls} (n={len(sub)})", alpha=0.6, s=25, edgecolors="none")

    ax1.set_xlabel("Heterozygosity (%)")
    ax1.set_ylabel("Distance to F5 Consensus")
    ax1.set_title("Het vs. Distance to F5")
    ax1.legend(fontsize=8)

    ax2.set_xlabel("Heterozygosity (%)")
    ax2.set_ylabel("Distance to Hembra Consensus")
    ax2.set_title("Het vs. Distance to Female Parent")
    ax2.legend(fontsize=8)

    fig.suptitle("Heterozygosity vs. Genetic Distance Metrics", fontsize=14, y=1.02)
    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig_ek4_distance_vs_het.png"), bbox_inches="tight")
    plt.close()


def fig_ek5_comparison(df, our_metrics, outdir):
    """Compare EK manual classification with our automated purity calls."""
    if our_metrics is None:
        return

    merged = df.merge(our_metrics[["sample", "purity_status", "het_rate", "concordance"]],
                      left_on="CELL", right_on="sample", how="left",
                      suffixes=("_ek", "_auto"))

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

    # Left: confusion matrix
    ct = pd.crosstab(merged["ek_class"], merged["purity_status"])
    for col in ["PASS", "WARNING", "FAIL"]:
        if col not in ct.columns:
            ct[col] = 0
    ct = ct[["PASS", "WARNING", "FAIL"]]

    im = ax1.imshow(ct.values, cmap="YlOrRd", aspect="auto")
    ax1.set_xticks(range(len(ct.columns)))
    ax1.set_xticklabels(ct.columns)
    ax1.set_yticks(range(len(ct.index)))
    ax1.set_yticklabels(ct.index)
    ax1.set_xlabel("Automated Purity Call")
    ax1.set_ylabel("EK Manual Classification")
    ax1.set_title("Classification Agreement")

    for i in range(len(ct.index)):
        for j in range(len(ct.columns)):
            val = ct.values[i, j]
            ax1.text(j, i, str(val), ha="center", va="center",
                     fontsize=14, fontweight="bold",
                     color="white" if val > ct.values.max() * 0.5 else "black")

    # Right: scatter comparing het rates
    for cls in ["Pure hybrid", "Self parent", "Outcross"]:
        sub = merged[merged["ek_class"] == cls]
        ax2.scatter(sub["het_rate_ek"] * 100, sub["concordance"] * 100 if "concordance" in sub.columns else 0,
                    c=STATUS_COLORS[cls], label=f"{cls} (n={len(sub)})",
                    alpha=0.6, s=25, edgecolors="none")

    ax2.set_xlabel("Heterozygosity (%, from EK sheet)")
    ax2.set_ylabel("Concordance with Group Consensus (%, automated)")
    ax2.set_title("EK Classification vs. Automated Concordance")
    ax2.legend(fontsize=8)

    plt.tight_layout()
    fig.savefig(os.path.join(outdir, "fig_ek5_comparison.png"), bbox_inches="tight")
    plt.close()


# ── README ─────────────────────────────────────────────────────────────────────
def build_ek_readme(df, lot_df):
    n_total = len(df)
    n_pure = (df["ek_class"] == "Pure hybrid").sum()
    n_self = (df["ek_class"] == "Self parent").sum()
    n_out = (df["ek_class"] == "Outcross").sum()

    lot_lines = []
    for _, row in lot_df.iterrows():
        lot_lines.append(
            f"  {row['lot']:10s}  n={row['n']:3.0f}  "
            f"self={row['self_rate']:.0%}  outcross={row['outcross_rate']:.0%}  "
            f"total_impurity={row['total_impurity']:.0%}"
        )

    readme = f"""README — EK Purity Analysis (Breeding Director's Manual Classification)
========================================================================

RELATIONSHIP TO RAW GENOTYPING DATA
This spreadsheet is a companion analysis to the raw genotyping data (260128_ZMS25508_B04_GT_Purity_data_Claude.xlsx). Both originate from the SAME sequencing run (ZMS25508 B04). This file covers {n_total} of the 538 total samples — specifically the F5 hybrid (n=233) and F5-DOBLE (n=40) samples. The breeding director performed a manual classification using genetic distance metrics.

DATA STRUCTURE
The "Pureza de F5" sheet contains one row per sample with:
  - CELL: Sample ID (plate-well format, e.g., "F6601 - P1A1")
  - GID / NAME: Germplasm identifier and genotype name
  - HIBRIDO/LINEA: All samples are "HIBRIDO" (hybrid)
  - Lote de Produccion: Seed production lot (L-99-1, L-99-2, L-88, L-117, L-VV3, R&D)
  - % Total Out-type/self: Lot-level impurity summary
  - % HETERO: Heterozygosity rate (%)
  - Self parent: "Yes" if the sample is classified as a selfed parent (not true hybrid)
  - Out-type/outcross: "Outcross" or "Yes" if the sample shows foreign pollen contamination
  - dist_prom_F5: Average genetic distance to F5 hybrid consensus genotype
  - dist_prom_hembra: Average genetic distance to Hembra (female parent) consensus
  - mas_cerca_de_hembra: Boolean — if True, sample is genetically closer to female parent than to F5 consensus (indicates selfing)

CLASSIFICATION METHODOLOGY
The breeding director classified samples based on genetic distance:
  - Self parents: Low heterozygosity (~12%), close to Hembra (female parent) genotype — these are seeds that resulted from self-pollination rather than cross-pollination with the male parent.
  - Outcross/out-type: Het rate similar to F5 but distant from the F5 consensus — these were pollinated by foreign pollen (not the intended male parent).
  - Pure hybrid: Normal het (~33%), close to F5 consensus — true hybrid seed.

RESULTS SUMMARY
  Pure hybrid:  {n_pure:3d} ({n_pure/n_total:.0%})
  Self parent:  {n_self:3d} ({n_self/n_total:.0%})
  Outcross:     {n_out:3d} ({n_out/n_total:.0%})

PER-LOT BREAKDOWN:
{chr(10).join(lot_lines)}

INTERPRETATION
- L-99-1 (Hamblin Duarte/Olopita) has the highest number of impurities: 20% selfed + 4% outcross. The breeding director notes this producer has had detasseling (desespigue) issues, and production area has already been reduced from 20 Mz to 7 Mz.
- L-88 (Eduardo Corado/Rancho) has 13% selfed + 17% outcross — the highest outcross rate, suggesting isolation distance issues.
- L-VV3 (Valle Verde) has 0% selfed but 10% outcross — good detasseling but potential isolation problems.
- R&D lot is nearly clean (2.5% total impurity).
- L-117 (Mynor Leiva) has moderate impurity at 13%.

The "Resumen OPS" sheet contains detailed field operations notes on detasseling compliance, plant atypicals, and production management actions.

FIGURES:
  fig_ek1: Distance scatter — F5 vs Hembra distance, colored by classification
  fig_ek2: Het rate histograms per classification
  fig_ek3: Per-lot classification stacked bar + impurity rate comparison
  fig_ek4: Het rate vs genetic distance scatter
  fig_ek5: Comparison of EK manual vs automated purity classification

HOW TO REPRODUCE:
  python3 ek_purity_analysis.py <path_to_ek_xlsx> [path_to_purity_summary.csv]
  Requires: pandas, numpy, matplotlib, openpyxl
"""
    return readme


def add_ek_readme_tab(xlsx_path, readme_text):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font

    wb = load_workbook(xlsx_path)
    if "README" in wb.sheetnames:
        del wb["README"]

    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 130
    ws["A1"] = readme_text
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A1"].font = Font(name="Consolas", size=10)
    ws.row_dimensions[1].height = 900

    wb.save(xlsx_path)
    print(f"  Added README tab to {os.path.basename(xlsx_path)}")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        candidates = [
            os.path.expanduser("~/Downloads/260128_ZMS25508_B04_GT_EK_F5 purity_12_02_2026.xlsx"),
        ]
        xlsx_path = None
        for c in candidates:
            if os.path.exists(c):
                xlsx_path = c
                break
        if xlsx_path is None:
            print("Usage: python3 ek_purity_analysis.py <path_to_ek_xlsx> [purity_summary.csv]")
            sys.exit(1)
    else:
        xlsx_path = sys.argv[1]

    our_csv = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "purity_summary.csv"
    )

    if not os.path.exists(xlsx_path):
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    outdir = os.path.dirname(os.path.abspath(__file__))
    print(f"Input:  {xlsx_path}")
    print(f"Output: {outdir}")
    print()

    print("Loading EK data...")
    df, summary = load_ek(xlsx_path)
    print(f"  {len(df)} samples, {df['Lote de Producción'].nunique()} lots")
    print(f"  Classification: {(df['ek_class'] == 'Pure hybrid').sum()} pure, "
          f"{(df['ek_class'] == 'Self parent').sum()} self, "
          f"{(df['ek_class'] == 'Outcross').sum()} outcross")

    our_metrics = load_our_analysis(our_csv)
    if our_metrics is not None:
        print(f"  Loaded automated analysis: {len(our_metrics)} samples")
    else:
        print(f"  No automated analysis found at {our_csv}")

    print()
    print("Quick summary (from breeding director):")
    print(summary.to_string(index=False))
    print()

    # Figures
    print("Generating figures...")
    fig_ek1_distance_scatter(df, outdir)
    print("  fig_ek1_distance_scatter.png")
    fig_ek2_het_by_classification(df, outdir)
    print("  fig_ek2_het_by_classification.png")
    lot_df = fig_ek3_lot_purity(df, outdir)
    print("  fig_ek3_lot_purity.png")
    fig_ek4_distance_vs_het(df, outdir)
    print("  fig_ek4_distance_vs_het.png")
    fig_ek5_comparison(df, our_metrics, outdir)
    print("  fig_ek5_comparison.png")

    # Summary CSV
    csv_path = os.path.join(outdir, "ek_purity_summary.csv")
    df[["CELL", "GID", "NAME", "Lote de Producción", "het_rate", "dist_F5",
        "dist_hembra", "closer_to_hembra", "ek_class"]].to_csv(csv_path, index=False)
    print(f"\nSaved: {csv_path}")

    # README tab
    print("\nUpdating workbook...")
    readme_text = build_ek_readme(df, lot_df)
    add_ek_readme_tab(xlsx_path, readme_text)

    print("\nDone!")


if __name__ == "__main__":
    main()
